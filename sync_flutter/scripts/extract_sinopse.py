#!/usr/bin/env python3
"""
Extract PUBLIC school enrollment data from INEP Sinopse Estatística.
Produces a compact JSON: {year: {ibge_code: {matriculas, eja, educacao_especial}}}
"""
import json
import os
import sys

import openpyxl


def safe_int(val):
    if val is None or val == "":
        return 0
    try:
        return int(float(str(val).strip().replace(".", "").replace(",", ".")))
    except (ValueError, TypeError):
        return 0


def extract_public_enrollments(xlsx_path: str, year: int) -> dict:
    """Extract rede pública enrollment from sheet 1.2."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {}

    # Sheet 1.2: Matrícula por localização e dependência
    ws = wb["1.2"]
    for row in ws.iter_rows(min_row=12, values_only=True):
        cod = row[3] if len(row) > 3 else None
        if not cod:
            continue
        cod_str = str(cod).strip()
        if not cod_str.isdigit() or len(cod_str) < 6:
            continue

        # Rede pública = Total - Privada(urbana) - Privada(rural)
        total = safe_int(row[4])
        privada_urb = safe_int(row[9])
        privada_rur = safe_int(row[14])
        publica = total - privada_urb - privada_rur

        # Estadual + Municipal breakdown
        estadual = safe_int(row[7]) + safe_int(row[12])
        municipal = safe_int(row[8]) + safe_int(row[13])

        if publica <= 0:
            continue

        result[cod_str] = {
            "m": publica,       # total matrículas públicas
            "e": estadual,      # estadual
            "mu": municipal,    # municipal
        }

    # Find EJA sheet dynamically (1.34 or 1.35 depending on year)
    eja_sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("eja") and "1." in name and "2." not in name and "3." not in name and "4." not in name:
            eja_sheet = name
            break
    if eja_sheet:
        try:
            ws_eja = wb[eja_sheet]
            for row in ws_eja.iter_rows(min_row=12, values_only=True):
                cod = row[3] if len(row) > 3 else None
                if not cod:
                    continue
                cod_str = str(cod).strip()
                if cod_str not in result:
                    continue
                eja_total = safe_int(row[4])
                if eja_total > 0:
                    result[cod_str]["ej"] = eja_total
        except Exception:
            pass

    # Find Educação Especial sheet dynamically (1.38 or 1.39)
    esp_sheet = None
    for name in wb.sheetnames:
        low = name.lower()
        if "especial" in low and "1." in name and "2." not in name and "3." not in name and "4." not in name:
            esp_sheet = name
            break
    if esp_sheet:
        try:
            ws_esp = wb[esp_sheet]
            for row in ws_esp.iter_rows(min_row=12, values_only=True):
                cod = row[3] if len(row) > 3 else None
                if not cod:
                    continue
                cod_str = str(cod).strip()
                if cod_str not in result:
                    continue
                esp_total = safe_int(row[4])
                if esp_total > 0:
                    result[cod_str]["es"] = esp_total
        except Exception:
            pass

    wb.close()
    return result


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sinopse_dir = os.path.join(base_dir, "sinopse_data")
    output_path = os.path.join(base_dir, "assets", "censo_matriculas.json")

    all_data = {}

    for root, dirs, files in os.walk(sinopse_dir):
        for f in files:
            if not f.endswith(".xlsx"):
                continue
            for y in range(2020, 2027):
                if str(y) in f:
                    xlsx_path = os.path.join(root, f)
                    print(f"Processing {y} from {f}...")
                    year_data = extract_public_enrollments(xlsx_path, y)
                    print(f"  Found {len(year_data)} municipalities")
                    all_data[str(y)] = year_data
                    break

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fp:
        json.dump(all_data, fp, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(output_path) / 1024
    print(f"\nSaved to {output_path} ({size_kb:.0f} KB)")
    print(f"Years: {sorted(all_data.keys())}")

    # Test
    for year in sorted(all_data.keys()):
        for city, code in [("Aracaju", "2800308"), ("Carambeí", "4104204")]:
            d = all_data[year].get(code, {})
            print(f"  {city} {year}: pub={d.get('m','-')}, eja={d.get('ej','-')}, esp={d.get('es','-')}")


if __name__ == "__main__":
    main()
